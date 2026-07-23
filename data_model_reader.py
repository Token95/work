import openpyxl
import os

# =====================================================================
# DATA MODEL READER
# Loads DataModel_Commands_Settings.xlsx and exposes lookup functions
# Used by cli_generator.py, generate_report.py, fieldwire_builder.py
# =====================================================================

DATAMODEL_PATH = "DataModel_Commands_Settings.xlsx"

# =====================================================================
# LOAD ALL SHEETS INTO MEMORY
# =====================================================================

def load_data_model(path=DATAMODEL_PATH):
    """
    Loads all three sheets from the Excel file.
    Returns a dict with keys: attributes, commands, device_config
    """
    if not os.path.exists(path):
        print(f"[-] DataModel not found at: {path}")
        return None

    wb = openpyxl.load_workbook(path, data_only=True)

    result = {
        "attributes":    [],
        "commands":      [],
        "device_config": {},
    }

    # --- SHEET 1: data model ---
    ws1 = wb["data model"]
    rows = list(ws1.iter_rows(values_only=True))

    for row in rows[1:]:
        attr_name = row[0]
        value     = row[1]
        cmd       = row[16] if len(row) > 16 else None

        if attr_name and cmd and str(cmd).startswith("zb"):
            result["attributes"].append({
                "name":    str(attr_name).strip(),
                "value":   value,
                "command": str(cmd).strip(),
            })

    # --- SHEET 3: Various Commands ---
    ws3 = wb["Various Commands"]
    rows3 = list(ws3.iter_rows(values_only=True))

    # Device config rows (Touch Combo, AIDA, ES Controller)
    for row in rows3:
        if not row[0]:
            continue
        device = str(row[0]).strip()
        if device in ["Touch Combo", "AIDA", "ES Controller"]:
            result["device_config"][device] = {
                "mqtt_server":   str(row[1]).strip() if row[1] else "",
                "mqtt_user":     str(row[2]).strip() if row[2] else "",
                "mqtt_pass":     str(row[3]).strip() if row[3] else "",
                "provision_key": str(row[4]).strip() if row[4] else "",
                "ota3_secret":   str(row[5]).strip() if row[5] else "",
            }

        # Host and ESP commands
        cmd_text = str(row[0]).strip()
        desc_text = str(row[1]).strip() if row[1] else ""
        if cmd_text and cmd_text not in ["Host commands", "ESP Command",
                                          "Variable field names", "Host Bridge Mode"]:
            result["commands"].append({
                "command":     cmd_text,
                "description": desc_text,
            })

    return result


# =====================================================================
# LOOKUP FUNCTIONS
# =====================================================================

def lookup_attribute(model, attr_name):
    """Find a specific attribute by name."""
    for attr in model["attributes"]:
        if attr["name"].lower() == attr_name.lower():
            return attr
    return None


def search_attributes(model, keyword):
    """Search attributes by keyword in name."""
    keyword = keyword.lower()
    return [a for a in model["attributes"] if keyword in a["name"].lower()]


def get_command(model, keyword):
    """Search commands by keyword in command text or description."""
    keyword = keyword.lower()
    return [c for c in model["commands"]
            if keyword in c["command"].lower()
            or keyword in c["description"].lower()]


def get_device_config(model, device_type):
    """Get MQTT and provision config for a device type."""
    return model["device_config"].get(device_type, {})


# =====================================================================
# SMARTCON STANDARD COMMISSIONING COMMANDS
# Built from data model lookups
# =====================================================================

def get_standard_commands(model, device_type, options):
    """
    Generate the full CLI command set for a device based on options.

    options = {
        "light_wiring":    "NC" or "NO",
        "light_terminal":  "W2" / "5" / etc,
        "reversing_valve": "B" or "O",
        "is_vrf":          True or False,
        "is_lg_vrf":       True or False,
        "fan_stages":      1 / 2 / 3,
        "set_auto":        True or False,
        "set_onoff":       True or False,
    }

    Returns list of (command, description) tuples.
    """
    cmds = []

    # 1. SET AUTO MODE
    if options.get("set_auto"):
        attr = lookup_attribute(model, "modes_available")
        if attr:
            cmds.append((attr["command"], "Set system to Auto mode"))
        else:
            cmds.append(("zb -a 0x0201 0xF040 2", "Set system to Auto mode"))

    # 2. ON/OFF BUTTON (AIDA and ES Controller only)
    if options.get("set_onoff") and device_type in ["AIDA", "ES Controller"]:
        cmds.append(("zb -a 0x0201 0xf092 1", "Set On/Off button to HVAC on/off"))

    # 3. REVERSING VALVE
    valve = options.get("reversing_valve", "O")
    attr = lookup_attribute(model, "changeover_relay_inverted")
    if valve == "B":
        cmds.append(("zb -a 0x0201 0xF090 1", "B-type reversing valve — inverted"))
    else:
        cmds.append(("zb -a 0x0201 0xF090 0", "O-type reversing valve — standard"))

    # 4. LIGHT RELAY INVERSION
    light_wiring  = options.get("light_wiring", "NO")
    light_terminal = options.get("light_terminal", "")

    AIDA_TERMINAL_MAP = {
        "2": 1, "3": 2, "5": 4, "6": 8, "7": 16, "9": 32
    }
    COMBO_TERMINAL_MAP = {
        "O": 1, "OB": 1, "G3": 1,
        "W2": 2, "*": 2,
        "G2": 4, "Y2": 4,
        "G": 8,
        "W": 16,
        "Y": 32,
    }

    if light_wiring == "NC" and light_terminal:
        t = light_terminal.upper().strip()
        if device_type in ["AIDA", "ES Controller"]:
            val = AIDA_TERMINAL_MAP.get(t, 0)
        else:
            val = COMBO_TERMINAL_MAP.get(t, 0)

        if val > 0:
            cmds.append((
                f"zb -a 0x0201 0xf09c {val}",
                f"NC wiring on terminal {t} — relay inverted"
            ))
        else:
            cmds.append((
                "zb -a 0x0201 0xf09c 0",
                f"Terminal {t} not mapped — no inversion"
            ))
    else:
        cmds.append((
            "zb -a 0x0201 0xf09c 0",
            "NO wiring — no relay inversion"
        ))

    # 5. VRF / LG VRF
    if options.get("is_vrf"):
        cmds.append((
            "zb -a 0x0201 0xf04f 1",
            "VRF interface — set primary temp input to Slot 1"
        ))

    if options.get("is_lg_vrf"):
        cmds.append((
            "zb -a 0x0201 0xf04f 1",
            "LG VRF interface — set primary temp input to Slot 1"
        ))
        fan_stages = options.get("fan_stages", 3)
        if device_type == "Touch Combo":
            cmds.append(("preset exec 45", "LG VRF preset — Touch Combo VRF only"))
        elif device_type == "AIDA":
            if fan_stages >= 3:
                cmds.append(("preset exec 91", "LG VRF preset — AIDA 3 fan"))
            else:
                cmds.append(("preset exec 92", "LG VRF preset — AIDA 2 fan"))
        elif device_type == "ES Controller":
            cmds.append(("preset exec 45", "LG VRF preset — ES Controller"))

    elif options.get("is_vrf"):
        fan_stages = options.get("fan_stages", 3)
        if device_type == "Touch Combo":
            cmds.append(("preset exec 45", "VRF preset — Touch Combo"))
        elif device_type == "AIDA":
            if fan_stages >= 3:
                cmds.append(("preset exec 91", "VRF preset — AIDA 3 fan"))
            else:
                cmds.append(("preset exec 92", "VRF preset — AIDA 2 fan"))
        elif device_type == "ES Controller":
            cmds.append(("preset exec 45", "VRF preset — ES Controller"))

    # 6. MQTT AND PROVISION
    config = get_device_config(model, device_type)
    if config:
        cmds.append((f"vs mqtt_server {config['mqtt_server']}", "Set MQTT server"))
        cmds.append((f"vs provision_key {config['provision_key']}", "Set provision key"))
        cmds.append((f"vs ota3_secret {config['ota3_secret']}", "Set OTA3 secret"))

    return cmds


# =====================================================================
# PRINT FUNCTIONS
# =====================================================================

def print_command_block(hotel, room_number, device_type, cmds):
    lines = []
    lines.append(f"# ============================================")
    lines.append(f"# SMARTCON CLI — {hotel} — Room {room_number}")
    lines.append(f"# Device: {device_type}")
    lines.append(f"# ============================================")
    for cmd, desc in cmds:
        lines.append(f"# {desc}")
        lines.append(cmd)
        lines.append("")
    lines.append(f"# ============================================")
    return lines


def save_command_block(lines, hotel, device_type):
    filename = f"CLI_{hotel.replace(' ', '_')[:25]}_{device_type.replace(' ', '_')}.txt"
    with open(filename, "w") as f:
        for line in lines:
            f.write(line + "\n")
    print(f"[+] CLI saved: {filename}")
    return filename


# =====================================================================
# STANDALONE TEST
# =====================================================================

if __name__ == "__main__":
    model = load_data_model()
    if not model:
        exit()

    print(f"\n[+] Loaded {len(model['attributes'])} attributes")
    print(f"[+] Loaded {len(model['commands'])} commands")
    print(f"[+] Device configs: {list(model['device_config'].keys())}")

    print("\n--- Searching for 'vrf' in attributes ---")
    vrf_attrs = search_attributes(model, "vrf")
    for a in vrf_attrs:
        print(f"  {a['name']:40} | {a['command']}")

    print("\n--- Searching for 'relay' in attributes ---")
    relay_attrs = search_attributes(model, "relay")
    for a in relay_attrs:
        print(f"  {a['name']:40} | {a['command']}")

    print("\n--- Auto mode command ---")
    auto = lookup_attribute(model, "modes_available")
    if auto:
        print(f"  {auto['name']} = {auto['command']}")

    print("\n--- Device configs ---")
    for device, config in model["device_config"].items():
        print(f"\n  {device}:")
        for k, v in config.items():
            print(f"    {k}: {v}")

    print("\n--- Sample CLI block for Touch Combo ---")
    cmds = get_standard_commands(model, "Touch Combo", {
        "set_auto":        True,
        "set_onoff":       False,
        "reversing_valve": "O",
        "light_wiring":    "NC",
        "light_terminal":  "W2",
        "is_vrf":          False,
        "is_lg_vrf":       False,
        "fan_stages":      3,
    })
    lines = print_command_block("Test Hotel", "101", "Touch Combo", cmds)
    for line in lines:
        print(line)